import os
import sys
import json
import zipfile
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Define color fills
HEADER_FILL = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Dark slate
PASS_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # Light green
FAIL_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Light red
KPI_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid") # Light gray

# Define fonts
FONT_TITLE = Font(name="Arial", size=16, bold=True, color="FFFFFF")
FONT_SECTION = Font(name="Arial", size=12, bold=True, color="0F172A")
FONT_HEADER = Font(name="Arial", size=10, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=9.5)
FONT_PASS = Font(name="Arial", size=9.5, bold=True, color="15803D")
FONT_FAIL = Font(name="Arial", size=9.5, bold=True, color="B91C1C")
FONT_KPI_VAL = Font(name="Arial", size=18, bold=True, color="1E293B")
FONT_KPI_LBL = Font(name="Arial", size=9, color="475569")

# Define alignments
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

# Define borders
BORDER_THIN = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0")
)

def create_summary_sheet(wb, title, total, passed, failed, modules_summary):
    ws = wb.active
    ws.title = "Summary"
    ws.views.sheetView[0].showGridLines = False

    # Title Banner
    ws.merge_cells("A1:G2")
    ws["A1"] = title
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 25

    # KPI Block Section
    ws["A4"] = "Executive Metrics"
    ws["A4"].font = FONT_SECTION

    # Counter Cards
    kpis = [
      ("Total Test Cases", total, "F1F5F9", "1E293B"),
      ("Passed Tests", passed, "DCFCE7", "15803D"),
      ("Failed Tests", failed, "FEE2E2", "B91C1C"),
      ("Pass Rate (%)", f"{(passed / total * 100):.1f}%" if total > 0 else "0.0%", "EFF6FF", "1D4ED8")
    ]

    for idx, (label, val, fill_color, font_color) in enumerate(kpis):
        col_start = 1 + (idx * 2)
        col_end = col_start + 1
        ws.merge_cells(start_row=5, start_column=col_start, end_row=5, end_column=col_end)
        ws.merge_cells(start_row=6, start_column=col_start, end_row=6, end_column=col_end)
        
        val_cell = ws.cell(row=5, column=col_start)
        val_cell.value = val
        val_cell.font = Font(name="Arial", size=18, bold=True, color=font_color)
        val_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        val_cell.alignment = Alignment(horizontal="center", vertical="bottom")

        lbl_cell = ws.cell(row=6, column=col_start)
        lbl_cell.value = label
        lbl_cell.font = Font(name="Arial", size=9, color=font_color)
        lbl_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        lbl_cell.alignment = Alignment(horizontal="center", vertical="top")

    ws.row_dimensions[5].height = 25
    ws.row_dimensions[6].height = 20

    # Module Distribution Table
    ws["A9"] = "Module Breakdown"
    ws["A9"].font = FONT_SECTION

    headers = ["Module / Feature Area", "Total Scans", "Passed", "Failed", "Pass Rate"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=10, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        cell.alignment = ALIGN_LEFT
        cell.border = BORDER_THIN
    ws.row_dimensions[10].height = 25

    curr_row = 11
    for mod_name, stats in modules_summary.items():
        ws.cell(row=curr_row, column=1, value=mod_name).alignment = ALIGN_LEFT
        ws.cell(row=curr_row, column=2, value=stats["total"]).alignment = ALIGN_RIGHT
        ws.cell(row=curr_row, column=3, value=stats["passed"]).alignment = ALIGN_RIGHT
        ws.cell(row=curr_row, column=4, value=stats["failed"]).alignment = ALIGN_RIGHT
        
        rate = (stats["passed"] / stats["total"] * 100) if stats["total"] > 0 else 0.0
        ws.cell(row=curr_row, column=5, value=f"{rate:.1f}%").alignment = ALIGN_RIGHT

        for c in range(1, 6):
            cell = ws.cell(row=curr_row, column=c)
            cell.font = FONT_BODY
            cell.border = BORDER_THIN
        ws.row_dimensions[curr_row].height = 20
        curr_row += 1

    # Adjust widths for summary sheet
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 22
    ws.column_dimensions["A"].width = 40

def style_data_sheet(ws):
    ws.views.sheetView[0].showGridLines = True
    ws.freeze_panes = "A2"
    
    # Style Header Row
    header_row = ws[1]
    ws.row_dimensions[1].height = 30
    for cell in header_row:
        cell.font = FONT_HEADER
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN

    # Add filters
    last_col = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

    # Style Body Rows
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 22
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=col_idx)
            cell.font = FONT_BODY
            cell.border = BORDER_THIN
            
            # Highlight Pass/Fail Status column
            if cell.value == "Pass" or cell.value == "PASS":
                cell.fill = PASS_FILL
                cell.font = FONT_PASS
                cell.alignment = ALIGN_CENTER
            elif cell.value == "Fail" or cell.value == "FAIL":
                cell.fill = FAIL_FILL
                cell.font = FONT_FAIL
                cell.alignment = ALIGN_CENTER
            elif col_idx in [1, 7, 8, 9, 10]:
                cell.alignment = ALIGN_CENTER
            else:
                cell.alignment = ALIGN_LEFT

    # Autofit Column Widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            # If multi-line, take longest line size
            lines = val.split('\n')
            for l in lines:
                if len(l) > max_len:
                    max_len = len(l)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

def generate_excel_report(json_path, excel_path, report_title):
    print(f"Reading results from {json_path}...")
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Test count validation
    total_cases = len(data)
    print(f"Loaded {total_cases} test cases.")
    if total_cases < 300:
        print(f"ERROR: Suite {report_title} contains only {total_cases} test cases! Minimum 300 required.")
        sys.exit(1)

    wb = openpyxl.Workbook()
    
    # 1. Executive Summary Sheet
    passed_cases = sum(1 for t in data if t.get("status") in ["Pass", "PASS"])
    failed_cases = total_cases - passed_cases

    modules_summary = {}
    for item in data:
        mod = item.get("module", "General")
        status = item.get("status", "Fail")
        if mod not in modules_summary:
            modules_summary[mod] = {"total": 0, "passed": 0, "failed": 0}
        modules_summary[mod]["total"] += 1
        if status in ["Pass", "PASS"]:
            modules_summary[mod]["passed"] += 1
        else:
            modules_summary[mod]["failed"] += 1

    create_summary_sheet(wb, report_title, total_cases, passed_cases, failed_cases, modules_summary)

    # 2. Detailed Test Results Sheet
    ws_details = wb.create_sheet(title="Test Results")
    
    # Header columns
    headers = [
        "Test Case ID", "Test Name", "Test Description", "Module", 
        "Input/Data", "Expected Result", "Actual Result", "Status", 
        "Execution Time", "Error/Failure Message", "Environment", "Timestamp"
    ]
    ws_details.append(headers)

    for item in data:
        row = [
            item.get("id", "N/A"),
            item.get("name", "N/A"),
            item.get("description", "N/A"),
            item.get("module", "N/A"),
            item.get("input", "N/A"),
            item.get("expected", "N/A"),
            item.get("actual", "N/A"),
            item.get("status", "Fail"),
            item.get("executionTime", "0ms"),
            item.get("error", ""),
            item.get("environment", "CI"),
            item.get("timestamp", "")
        ]
        ws_details.append(row)

    style_data_sheet(ws_details)

    wb.save(excel_path)
    print(f"Saved Excel report to {excel_path}")
    return total_cases, passed_cases, failed_cases

def create_zip(zip_path, file_to_zip, folder_in_zip=None):
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        if folder_in_zip:
            arcname = os.path.join(folder_in_zip, os.path.basename(file_to_zip))
        else:
            arcname = os.path.basename(file_to_zip)
        zipf.write(file_to_zip, arcname)
    print(f"Created ZIP archive: {zip_path}")

def main():
    root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    
    suites = [
        (os.path.join("selenium-tests", "selenium-results.json"), "selenium-report.xlsx", "selenium-report.zip", "Selenium UI Testing", "Selenium"),
        (os.path.join("appium-tests", "appium-results.json"), "appium-report.xlsx", "appium-report.zip", "Appium Mobile Testing", "Appium"),
        (os.path.join("validation-tests", "validation-results.json"), "field-validation-report.xlsx", "field-validation-report.zip", "Form Field Validation Testing", "Field-Validation"),
        (os.path.join("security-tests", "security-results.json"), "vulnerability-report.xlsx", "vulnerability-report.zip", "Security & Vulnerability Testing", "Vulnerability"),
        ("load-results.json", "load-test-report.xlsx", "load-test-report.zip", "Load & Performance Testing", "Load-Test")
    ]

    summary_stats = []

    print("--- Starting Excel and ZIP Report Generation ---")

    for results_json, report_xlsx, report_zip, title, subfolder in suites:
        json_path = os.path.join(root, results_json)
        xlsx_path = os.path.join(root, report_xlsx)
        zip_path = os.path.join(root, report_zip)

        if not os.path.exists(json_path):
            print(f"CRITICAL ERROR: Results file not found: {json_path}")
            sys.exit(1)

        total, passed, failed = generate_excel_report(json_path, xlsx_path, title)
        create_zip(zip_path, xlsx_path)
        summary_stats.append((title, total, passed, failed, xlsx_path, subfolder))

    # Create Master Package all-test-reports.zip
    master_zip_path = os.path.join(root, "all-test-reports.zip")
    print(f"Creating master reports package: {master_zip_path}...")
    with zipfile.ZipFile(master_zip_path, 'w', zipfile.ZIP_DEFLATED) as master_zip:
        for title, total, passed, failed, xlsx_path, subfolder in summary_stats:
            arcname = os.path.join(subfolder, os.path.basename(xlsx_path))
            master_zip.write(xlsx_path, arcname)
    print(f"Master reports package created successfully.")

    # Write GitHub Step Summary
    summary_markdown = "## E2E Test Execution Summary (1500+ Test Cases)\n\n"
    summary_markdown += "| Test Category | Executed Cases | Passed | Failed | Pass Rate |\n"
    summary_markdown += "| :--- | :---: | :---: | :---: | :---: |\n"

    for title, total, passed, failed, _, _ in summary_stats:
        rate = (passed / total * 100) if total > 0 else 0.0
        summary_markdown += f"| **{title}** | {total} | {passed} | {failed} | {rate:.1f}% |\n"

    summary_markdown += "\n- Reports have been packaged as separate zip folders in workflow artifacts."

    step_summary_env = os.environ.get("GITHUB_STEP_SUMMARY")
    if step_summary_env:
        with open(step_summary_env, "w", encoding="utf-8") as f:
            f.write(summary_markdown)
        print("Written GitHub Actions Step Summary.")
    else:
        print("\n" + summary_markdown + "\n")

    print("--- Report Generation Completed Successfully ---")

if __name__ == "__main__":
    main()
